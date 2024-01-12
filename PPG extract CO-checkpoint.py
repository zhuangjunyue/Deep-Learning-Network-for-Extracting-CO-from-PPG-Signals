#!/usr/bin/env python
# coding: utf-8

# 2024/1/9 这是提取PPG信号中各个参数并进行回归拟合心输出量的代码

# 首先对PPG数据进行解包绘图 处理PPG数据时只需改变file_path中文件名 由于数据量过多 分六个子图绘制 此时数据存储在original_ppg中

# In[2048]:


import matplotlib.pyplot as plt
import pandas as pd

# 初始化变量
original_ppg = []
loaded_successfully = False
error_message = ''

try:
    # 使用pandas直接读取第一列数据
    #file_path = 'path_to_your_file.txt'  # 指定文件路径
    data = pd.read_csv(file_path, sep='\t', header=None, usecols=[0], skiprows=2)
    original_ppg = data[0].tolist()

    # 设置采样率等参数
    sampling_rate = 500  # Hz
    cycle_duration = 1  
    points_per_cycle = 500

    # 计算分割段
    cycles_per_segment = 20
    segments = [original_ppg[i * points_per_cycle * cycles_per_segment:(i + 1) * points_per_cycle * cycles_per_segment] for i in range(6)]

    # 绘图
    fig, axs = plt.subplots(6, 1, figsize=(10, 20))
    for i, segment in enumerate(segments):
        axs[i].plot(segment)
        axs[i].set_title(f'PPG data {i+1} with 20 Cycles')
        axs[i].set_xlabel('Sample Number')
        axs[i].set_ylabel('PPG Amplitude')

    plt.tight_layout()
    plt.show()

    loaded_successfully = True
except Exception as e:
    loaded_successfully = False
    error_message = str(e)

# 输出是否成功加载和错误消息
print("Loaded Successfully:", loaded_successfully)
if not loaded_successfully:
    print("Error Message:", error_message)


# 以上六个子图是对解包后的original_ppg数据绘图 观察各个波段PPG质量好坏 后续分析中需选择质量好的PPG数据进行分析

# 接下来选择截取质量好的PPG波段数据 将这一段数据进行后续分析 此时数据存储在 ppg_data中

# In[2049]:


ppg_data = original_ppg[0:50000]
# 计算分割段[a:b] a为数据开始的索引 b为数据结束的索引
cycles_per_segment = 10
segments = [ppg_data[i * points_per_cycle * cycles_per_segment:(i + 1) * points_per_cycle * cycles_per_segment] for i in range(10)]

# 绘图
fig, axs = plt.subplots(10, 1, figsize=(10, 20))
for i, segment in enumerate(segments):
    axs[i].plot(segment)
    axs[i].set_title(f'PPG data {i+1} with 20 Cycles')
    axs[i].set_xlabel('Sample Number')
    axs[i].set_ylabel('PPG Amplitude')

plt.tight_layout()
plt.show()


# 以上六个子图是选择数据质量较好的PPG波段进行绘制得到的子图  ppg_data

# 对PPG信号设置巴特沃斯3阶带通滤波器 通带1HZ到15HZ 绘制滤波后的PPG波形图与滤波器频率响应图 此时波形存储在filtered_ppg

# In[2050]:


import numpy as np
import matplotlib.pyplot as plt
from scipy.signal import butter, filtfilt

# 高通和低通滤波器的参数设置
lowcut = 1  # 低通滤波器截止频率 HZ
highcut = 15  # 高通滤波器截止频率 HZ
sampling_rate = 500  # 采样率，单位为Hz
order = 3  # 滤波器的阶数 不能过高 过高把大部分频率信息都丢失

# Butterworth滤波器函数 函数内参 lowcut低频截止频率 highcut高频截止频率 fs采样率 order结束 函数内参可直接调整传入函数
def butter_bandpass(lowcut, highcut, fs, order=3):
    nyq = 0.5 * fs
    low = lowcut / nyq
    high = highcut / nyq
    b, a = butter(order, [low, high], btype='band')
    return b, a

# 使用Butterworth滤波器对数据进行滤波的函数 data传入函数需要调整的参数 lowcut highcut高低截止频率 fs采样率 order滤波器阶数
def butter_bandpass_filter(data, lowcut, highcut, fs, order=3):
    b, a = butter_bandpass(lowcut, highcut, fs, order=order)
    y = filtfilt(b, a, data)
    return y

# 对PPG数据进行滤波 滤波后数据存储入filtered_ppg 
filtered_ppg = butter_bandpass_filter(ppg_data, lowcut, highcut, sampling_rate, order)

ppg_spectrum = np.fft.rfft(filtered_ppg)
frequencies = np.fft.rfftfreq(len(filtered_ppg), 1/sampling_rate)
# 绘制滤波后的PPG图像
ppg_spectrum = np.fft.rfft(filtered_ppg)
frequencies = np.fft.rfftfreq(len(filtered_ppg), 1/sampling_rate)
# 计算频谱的幅度
magnitude = np.abs(ppg_spectrum)

# 绘制频谱图
plt.figure(figsize=(10, 6))
plt.plot(frequencies, magnitude)
plt.title('PPG Frequency Spectrum')
plt.xlabel('Frequency (Hz)')
plt.ylabel('Magnitude')
plt.xlim(0, 15)  # 展示 0-5Hz 的频谱范围 正常PPG波的基本频率应在1Hz（60次/分钟）到1.67Hz（100次/分钟）之间 低频包含的成分一般为呼吸成分
plt.grid(True)
plt.show()
# 分6个子图绘制滤波后的PPG图像
fig, axs = plt.subplots(6, 1, figsize=(10, 20))
points_per_segment = len(ppg_data) // 6
print(points_per_segment)
print(len(ppg_data))

for i in range(6):
    start_index = i * points_per_segment
    end_index = start_index + points_per_segment
    axs[i].plot(filtered_ppg[start_index:end_index])
    axs[i].set_title(f'Filtered PPG Segment {i+1}')
    axs[i].set_xlabel('Sample Number')
    axs[i].set_ylabel('Amplitude')

plt.tight_layout()
plt.show()

import numpy as np
import matplotlib.pyplot as plt
from scipy.signal import butter, filtfilt, freqz

# 参数设置
lowcut = 1  # 低通滤波器截止频率 1 Hz
highcut = 15  # 高通滤波器截止频率 15 Hz
sampling_rate = 500  # 采样率，单位为Hz
order = 3  # 滤波器的阶数

# 巴特沃斯带通滤波器设计
def butter_bandpass(lowcut, highcut, fs, order=3):
    nyq = 0.5 * fs
    low = lowcut / nyq
    high = highcut / nyq
    b, a = butter(order, [low, high], btype='band')
    return b, a

# 定义滤波函数
def butter_bandpass_filter(data, lowcut, highcut, fs, order=3):
    b, a = butter_bandpass(lowcut, highcut, fs, order=order)
    y = filtfilt(b, a, data)
    return y

# 设计滤波器并计算频率响应
b, a = butter_bandpass(lowcut, highcut, sampling_rate, order)
w, h = freqz(b, a, worN=8000)

# 绘制滤波器的频率响应图
plt.figure(figsize=(10, 6))
plt.plot(0.5 * sampling_rate * w / np.pi, np.abs(h), 'b')
plt.plot(lowcut, 0.5*np.sqrt(2), 'ko')
plt.plot(highcut, 0.5*np.sqrt(2), 'ko')
plt.axvline(lowcut, color='k')
plt.axvline(highcut, color='k')
plt.xlim(0, 0.5 * sampling_rate)
plt.title("Butterworth Bandpass Filter Frequency Response")
plt.xlabel('Frequency [Hz]')
plt.ylabel('Gain')
plt.grid()

# 计算纹波抑制比（在截止频率处）
ripple_suppression = 20 * np.log10(np.min(np.abs(h)))
print(f"Ripple Suppression (dB): {ripple_suppression:.2f}")

plt.show()


# 上图是PPG波滤波后绘制 filtered_ppg 以及滤波后频谱图 峰值代表心率频率 次峰值一般代表二尖瓣波值频率 其他高阶高频成分 

# PPG波中低频的成分表征基线漂移 也表征呼吸成分 使用巴特斯沃滤波器提取 此时实验体使用呼吸机非自主呼吸 提取出数据存储在ppg_breathing中

# In[2051]:


#提取低频基线漂移表征呼吸的成分 进行低通滤波
# 设置低通滤波器的参数
lowcut_breathing = 0.5  # 低通滤波器截止频率，单位为Hz

# Butterworth滤波器设计函数，用于低通滤波
def butter_lowpass(cutoff, fs, order=5):
    nyq = 0.5 * fs
    normal_cutoff = cutoff / nyq
    b, a = butter(order, normal_cutoff, btype='low', analog=False)
    return b, a

# 使用Butterworth滤波器对数据进行低通滤波的函数
def butter_lowpass_filter(data, cutoff, fs, order=5):
    b, a = butter_lowpass(cutoff, fs, order=order)
    y = filtfilt(b, a, data)
    return y

# 对PPG数据进行低通滤波以提取呼吸成分
ppg_breathing = butter_lowpass_filter(ppg_data, lowcut_breathing, sampling_rate, order)

# 绘制滤波后的PPG呼吸成分图像
plt.figure(figsize=(10, 6))
plt.plot(ppg_breathing)
plt.title('PPG Breathing Component (0-0.5 Hz)')

plt.xlabel('Sample Number')
plt.ylabel('Amplitude')
plt.show()





# 上图是PPG中低频呼吸成分 基线漂移 周期一般代表呼吸分量

# 使用自相关函数方法 提出质量不好的PPG波形 自相关波形的峰值若有明显周期性 则说明波段周期性良好 设置不好的波形在自相关函数的第一个峰值之前 绘制自相关图以及剔除不好波形后的PPG 此时数据存储在ppg_filtered_new

# In[2052]:


import numpy as np
import matplotlib.pyplot as plt
from scipy.signal import find_peaks

ppg_filtered = filtered_ppg

# 计算自相关函数
autocorr = np.correlate(ppg_filtered, ppg_filtered, mode='full')
autocorr = autocorr[autocorr.size // 2:]  # 取自相关的一半

# 检测波形质量 使用自相关的峰值
# 找到除零延迟之外的峰值
peaks, _ = find_peaks(autocorr, distance=250)  # distance参数根据PPG信号的特性调整

# 如果找到多个峰值，考虑第一个非零延迟峰值
if len(peaks) > 1:
    quality_peak = peaks[1]  # 第一个非零延迟峰值
    ppg_filtered_new = ppg_filtered[quality_peak:]
else:
    print("未找到合适的峰值，保留原始数据。")
    ppg_filtered_new = ppg_filtered
    quality_peak = 0

# 绘制自相关图
plt.figure(figsize=(10, 4))
plt.plot(autocorr)
plt.title('PPG self correlate')
plt.xlabel('below')
plt.ylabel('correlation')
plt.show()

# 将ppg_filtered分为6个子图绘制
subplots = 6
samples_per_subplot = len(ppg_filtered) // subplots

fig, axes = plt.subplots(subplots, 1, figsize=(10, 12))
for i in range(subplots):
    axes[i].plot(ppg_filtered[i * samples_per_subplot: (i + 1) * samples_per_subplot])
    axes[i].set_title(f'ppg_filtered  {i+1}')
    axes[i].set_xlabel('index')
    axes[i].set_ylabel('PPG altitude')
plt.tight_layout()
plt.show()

# 将ppg_filtered_new分为6个子图绘制，并标注剔除的部分
samples_per_subplot_new = len(ppg_filtered_new) // subplots

fig, axes = plt.subplots(subplots, 1, figsize=(10, 12))
for i in range(subplots):
    start_index = i * samples_per_subplot_new
    end_index = (i + 1) * samples_per_subplot_new
    axes[i].plot(ppg_filtered_new[start_index: end_index])
    axes[i].set_title(f'ppg_filtered_new  {i+1}')
    axes[i].set_xlabel('index')
    axes[i].set_ylabel('PPG altitude')
    
plt.tight_layout()
plt.show()


# 上图是自相关波形以及ppg_filtered_new波形的绘制

# 在滤波后提取心率与呼吸成分 使用find peaks峰值谷值检测方法 初步确定心跳与呼吸周期 此时对于峰值谷值的标注是有问题 会把局部最小值标注为峰值谷值 

# In[2053]:


from scipy.signal import find_peaks

# ppg_breathing 和 filtered_ppg 滤波之后存储用于表征PPG正常波形反应心率以及低频反映呼吸
# PPG呼吸信号的峰值和谷值 后面发现find peak这个python内置的函数会把所有一阶导数为零点的值识别为峰值谷值 为修改 后面设置峰峰值之间最小间距
peaks_breathing, _ = find_peaks(ppg_breathing)
troughs_breathing, _ = find_peaks(-ppg_breathing)

# 找到PPG心率信号的峰值和谷值
peaks_heartbeat, _ = find_peaks(ppg_filtered_new)
troughs_heartbeat, _ = find_peaks(-ppg_filtered_new)

# 计算呼吸周期的均值 均值作为后面峰峰值间距最小值的依据
breathing_intervals = np.diff(peaks_breathing)
average_breathing_interval = np.mean(breathing_intervals) / sampling_rate

# 计算心率周期的均值
heartbeat_intervals = np.diff(peaks_heartbeat)
average_heartbeat_interval = np.mean(heartbeat_intervals) / sampling_rate

# 分为6个子图绘制每个信号的峰值和谷值
points_per_segment_breathing = len(ppg_breathing) // 6
points_per_segment_heartbeat = len(ppg_filtered_new) // 6

fig, axs = plt.subplots(12, 1, figsize=(12, 24))

for i in range(6):
    # 呼吸信号的子图
    start_index_breathing = i * points_per_segment_breathing
    end_index_breathing = start_index_breathing + points_per_segment_breathing
    segment_breathing = ppg_breathing[start_index_breathing:end_index_breathing]
    peaks_segment_breathing = peaks_breathing[(peaks_breathing >= start_index_breathing) & (peaks_breathing < end_index_breathing)] - start_index_breathing
    troughs_segment_breathing = troughs_breathing[(troughs_breathing >= start_index_breathing) & (troughs_breathing < end_index_breathing)] - start_index_breathing

    axs[i].plot(segment_breathing, label='Breathing Signal')
    axs[i].plot(peaks_segment_breathing, segment_breathing[peaks_segment_breathing], 'x', label='Peaks')
    axs[i].plot(troughs_segment_breathing, segment_breathing[troughs_segment_breathing], 'o', label='Troughs')
    axs[i].set_title(f'PPG Breathing Segment {i+1}')
    axs[i].legend()

    # 心率信号的子图
    start_index_heartbeat = i * points_per_segment_heartbeat
    end_index_heartbeat = start_index_heartbeat + points_per_segment_heartbeat
    segment_heartbeat = filtered_ppg[start_index_heartbeat:end_index_heartbeat]
    peaks_segment_heartbeat = peaks_heartbeat[(peaks_heartbeat >= start_index_heartbeat) & (peaks_heartbeat < end_index_heartbeat)] - start_index_heartbeat
    troughs_segment_heartbeat = troughs_heartbeat[(troughs_heartbeat >= start_index_heartbeat) & (troughs_heartbeat < end_index_heartbeat)] - start_index_heartbeat

    axs[i + 6].plot(segment_heartbeat, label='Heartbeat Signal')
    axs[i + 6].plot(peaks_segment_heartbeat, segment_heartbeat[peaks_segment_heartbeat], 'x', label='Peaks')
    axs[i + 6].plot(troughs_segment_heartbeat, segment_heartbeat[troughs_segment_heartbeat], 'o', label='Troughs')
    axs[i + 6].set_title(f'PPG Heartbeat Segment {i+1}')
    axs[i + 6].legend()

plt.tight_layout()
plt.show()

average_breathing_interval, average_heartbeat_interval
#输出平均周期


# 以上是对ppg_filtered_new初步标注峰值谷值的图像 输出初步周期 依据初步周期设置峰峰值最小间隔参数 distance参数 避免将局部最小值错误标注为峰值谷值 distance参数依据绘图后结果调整 调整后重新标注峰值谷值并绘图 提取呼吸周期与呼吸周期变异率

# In[2054]:


# 为确保每个周期内只有一个峰值和一个谷值，我们将调整峰值和谷值的提取方法
# 假定每个呼吸周期至少持续2秒钟基于正常的呼吸频率

min_distance_between_peaks = int(sampling_rate * 2)  # 最小峰值间距设为4秒

# 使用find_peaks函数，增加distance参数以确保每个周期内只检测到一个峰值和谷值
peaks_breathing, _ = find_peaks(ppg_breathing, distance=min_distance_between_peaks)
troughs_breathing, _ = find_peaks(-ppg_breathing, distance=min_distance_between_peaks)
breathing_intervals = np.diff(peaks_breathing)
average_breathing_intervalnew = np.mean(breathing_intervals) / sampling_rate #呼吸平均周期
breathing_periods = np.diff(peaks_breathing) / sampling_rate  
breathing_period_variability = np.std(breathing_periods) / np.mean(breathing_periods) #呼吸周期变异率
# 分6个子图绘制调整后的峰值和谷值
# 修改峰值和谷值索引的处理方式，以避免索引越界的错误

fig, axs = plt.subplots(6, 1, figsize=(12, 18))

for i in range(6):
    start_index = i * points_per_segment_breathing
    end_index = start_index + points_per_segment_breathing
    segment_breathing = ppg_breathing[start_index:end_index]

    # 确保峰值和谷值索引在当前分段的范围内
    segment_peaks = [peak for peak in peaks_breathing if start_index <= peak < end_index]
    segment_troughs = [trough for trough in troughs_breathing if start_index <= trough < end_index]

    # 将全局索引转换为当前分段的相对索引
    segment_peaks_relative = [peak - start_index for peak in segment_peaks]
    segment_troughs_relative = [trough - start_index for trough in segment_troughs]

    axs[i].plot(segment_breathing, label='Breathing Signal')
    axs[i].plot(segment_peaks_relative, segment_breathing[segment_peaks_relative], 'x', label='Peaks')
    axs[i].plot(segment_troughs_relative, segment_breathing[segment_troughs_relative], 'o', label='Troughs')
    axs[i].set_title(f'PPG Breathing Segment {i+1}')
    axs[i].legend()

plt.tight_layout()
plt.show()
print(average_breathing_intervalnew)
print(breathing_period_variability)
print(breathing_periods)
print(breathing_intervals)
#输出呼吸周期均值 呼吸周期变异率


# 对 0.5HZ-15HZ PPG波形做相同处理 distance参数可调 重新标注峰值谷值并绘图 提取精确周期与心率变异率

# In[2055]:


# 为确保每个周期内只有一个峰值和一个谷值，我们将调整峰值和谷值的提取方法
# 假定每个PPG周期至少持续0.5秒钟
ppg_filtered=ppg_filtered_new
points_per_segment_filtered=points_per_segment_breathing
min_distance_between_peaks = int(sampling_rate * 0.35)  # 最小峰值间距设为0.5秒

# 使用find_peaks函数，增加distance参数以确保每个周期内只检测到一个峰值和谷值
peaks_filtered, _ = find_peaks(ppg_filtered, distance=min_distance_between_peaks)
troughs_filtered, _ = find_peaks(-ppg_filtered, distance=min_distance_between_peaks)
filtered_intervals = np.diff(peaks_filtered)
average_filtered_interval = np.mean(filtered_intervals) / sampling_rate #呼吸平均周期
filtered_periods = np.diff(peaks_filtered) / sampling_rate  
filtered_period_variability = np.std(filtered_periods) / np.mean(filtered_periods) #呼吸周期变异率
# 分6个子图绘制调整后的峰值和谷值
# 修改峰值和谷值索引的处理方式，以避免索引越界的错误

fig, axs = plt.subplots(6, 1, figsize=(12, 18))

for i in range(6):
    start_index = i * points_per_segment_filtered
    end_index = start_index + points_per_segment_filtered
    segment_filtered = ppg_filtered[start_index:end_index]

    # 确保峰值和谷值索引在当前分段的范围内
    segment_peaks = [peak for peak in peaks_filtered if start_index <= peak < end_index]
    segment_troughs = [trough for trough in troughs_filtered if start_index <= trough < end_index]

    # 将全局索引转换为当前分段的相对索引
    segment_peaks_relative = [peak - start_index for peak in segment_peaks]
    segment_troughs_relative = [trough - start_index for trough in segment_troughs]

    axs[i].plot(segment_filtered, label='filtered Signal')
    axs[i].plot(segment_peaks_relative, segment_filtered[segment_peaks_relative], 'x', label='Peaks')
    axs[i].plot(segment_troughs_relative, segment_filtered[segment_troughs_relative], 'o', label='Troughs')
    axs[i].set_title(f'PPG filtered Segment {i+1}')
    axs[i].legend()

plt.tight_layout()
plt.show()
print(average_filtered_interval)
print(filtered_period_variability)
print(filtered_periods)
print(filtered_intervals)


# 提取PPG波的直流分量与交流分量 先提取A1表征交流分量的峰峰值 A2表征二尖瓣降波值（峰值之后第一个曲率为零的点与谷值之间垂直距离）A3表征直流分量 对于A3的提取使用原始数据中DC分量直流均值 对于A1,A2的提取 由于有的PPG波中不存在二尖瓣波值 需选择完好的波形

# 首先提取DC分量 用orinal_ppg的数据均值作为直流分量

# In[2056]:


# 提取DC分量 - 使用原始PPG数据的直流均值 原始没有经过滤波的PPG数据中包含原始DC分量信息 滤波去除低频分量会滤除DC信息
DC_value = np.mean(ppg_data)
print(DC_value)


# 提取交流振幅分量 两个波谷界定一个周期 周期中波峰与较低波谷的垂直距离作为振幅分量

# In[2057]:


import numpy as np
import matplotlib.pyplot as plt
from scipy.signal import find_peaks

# 设置最小峰值和谷值间隔，一个周期至少持续0.5秒，采样率为500Hz
min_peak_distance = 180  # 0.5秒对应的采样点数

# 使用find_peaks找到满足条件的峰值
peaks, _ = find_peaks(ppg_filtered, distance=min_peak_distance)

# 使用find_peaks找到满足条件的谷值
troughs, _ = find_peaks(-ppg_filtered, distance=min_peak_distance)

# 计算每个峰值和相邻谷值之间的差值（峰峰值）
peak_to_trough_amplitudes = [ppg_filtered[peak] - ppg_filtered[troughs[np.argmax(troughs > peak) - 1]] for peak in peaks if np.argmax(troughs > peak) > 0]

# 计算A1，即峰峰值的平均值
Altitude = np.mean(peak_to_trough_amplitudes) if peak_to_trough_amplitudes else 0

fig, axs = plt.subplots(6, 1, figsize=(12, 18))
segment_length = 10000
for i in range(6):
    start_index = i * segment_length
    end_index = start_index + segment_length
    segment = ppg_filtered[start_index:end_index]

    # 调整峰值和谷值的索引，确保它们位于当前分段内
    segment_peaks = [peak - start_index for peak in peaks if start_index <= peak < end_index]
    segment_troughs = [trough - start_index for trough in troughs if start_index <= trough < end_index]

    axs[i].plot(segment, label='Filtered PPG Segment')
    axs[i].plot(segment_peaks, segment[segment_peaks], 'x', label='Peaks', color='red')
    axs[i].plot(segment_troughs, segment[segment_troughs], 'x', label='Troughs', color='green')
    axs[i].set_title(f'Segment {i+1}')
    axs[i].legend()

plt.tight_layout()
plt.show()

# 输出A1的值
print(Altitude)
print(peak_to_trough_amplitudes)


# 振幅随时间呈现规律性变化 反映低频心脏的泵血效率 振幅变化周期提取出来 (其实是呼吸的基频）

# In[2058]:


from scipy.signal import correlate, find_peaks
import matplotlib.pyplot as plt

# 计算自相关
auto_corr = correlate(peak_to_trough_amplitudes, peak_to_trough_amplitudes, mode='full')
center = len(auto_corr) // 2

# 从自相关的中心开始寻找第一个峰值，这个峰值代表了周期的长度
auto_corr_peaks, _ = find_peaks(auto_corr[center:])

# 检查是否找到显著的峰值
if len(auto_corr_peaks) > 10:
    # 第一个峰值的位置即为周期
    period = auto_corr_peaks[0]
else:
    # 没有找到显著的峰值，将周期设置为0
    period = 0

# 根据周期计算其他相关参数
period_true = period * 0.57
print(period_true)

# 绘制自相关结果
plt.figure(figsize=(12, 6))
plt.plot(auto_corr[center:center + 2 * period] if period > 0 else auto_corr[center:])
plt.title('Autocorrelation of Peak-to-Trough Amplitudes')
plt.xlabel('Lag')
plt.ylabel('Autocorrelation')
plt.grid()
plt.show()

# 绘制原始数据
plt.figure(figsize=(12, 6))
plt.plot(peak_to_trough_amplitudes)
plt.title('Peak-to-Trough Amplitudes Over Time')
plt.xlabel('Data Point Index')
plt.ylabel('Amplitude')
plt.grid()
plt.show()


# 提取二尖瓣波值 方法：首先先将峰值与谷值在图中标注出来，然后寻找峰值与下一个峰值之间所有局部最小值点（一阶导数为零，并且这个点要排除掉谷值）然后输出并比较他们二阶导数的大小，最后选择二阶导最大的点标注出来 然后增加约束条件 这个点与周期内峰值的垂直距离大于1/3振幅 与周期内谷值距离大于1/3振幅 选择列表中符合条件且一阶导数为0二阶导数最大的点作为二尖瓣波值点

# 先观察 选择有明显二尖瓣波值的波段存储在ppg_segment中并绘图

# In[2059]:


ppg_segment = ppg_filtered_new[0:10000]

# Plot the segment
plt.figure(figsize=(10, 4))
plt.plot(ppg_segment)
plt.title('PPG Signal Segment (3000th to 4000th index)')
plt.xlabel('Sample Index')
plt.ylabel('PPG Amplitude')
plt.grid(True)
plt.show()


# 开始标注二尖瓣波值 标注好后绘图 输出二尖瓣波值与周期内谷值距离列表

# In[2064]:


import numpy as np
from scipy.signal import find_peaks, argrelextrema

import numpy as np
from scipy.signal import find_peaks, argrelextrema
#ppg_segment = np.array(ppg_data)
#赋值 列表转化为数组


    
def find_dicrotic_notches_corrected(signal):
    # 找到所有的峰值和谷值，峰值之间的最小间隔设为190个采样点，谷值间隔设为190
    peaks, _ = find_peaks(signal, distance=200)
    troughs, _ = find_peaks(-signal, distance=200)

    # 计算信号的一阶和二阶导数
    first_derivative = np.diff(signal)
    second_derivative = np.diff(first_derivative)
    
    # 初始化二尖瓣波值列表
    dicrotic_notches = []

    # 遍历每个峰值，寻找它和下一个峰值之间的所有局部最小值点
    for i in range(len(peaks) - 1):
        # 检查i+1是否在troughs数组的索引范围内
        if i + 1 < len(troughs):
            # 在当前峰值和下一个峰值之间找到所有的局部最小值点
            local_minima = argrelextrema(signal[peaks[i]:peaks[i+1]], np.less)[0] + peaks[i]
            # 排除谷值
            local_minima = [index for index in local_minima if index not in troughs]

            # 初始化用于比较的二阶导数最大值
            max_second_derivative = -np.inf
            # 初始化用于记录当前最优的二尖瓣波值索引
            best_notch = None

            # 遍历局部最小值点，找出二阶导数最大且符合振幅条件的点
            for local_minimum in local_minima:
                # 确保local_minimum不超出signal的索引范围
                if local_minimum < len(signal) and local_minimum - 1 < len(second_derivative):
                    # 检查与波峰和波谷的振幅条件
                    amplitude_condition_peak = abs(signal[peaks[i]] - signal[local_minimum]) > (abs(signal[peaks[i]] - signal[troughs[i]]) / 10)
                    amplitude_condition_trough = abs(signal[troughs[i+1]] - signal[local_minimum]) > (abs(signal[troughs[i+1]] - signal[peaks[i+1]]) / 10)
                    amplitude_condition = amplitude_condition_peak and amplitude_condition_trough

                    # 检查二阶导数的大小，并更新最优值
                    if second_derivative[local_minimum - 1] > max_second_derivative and amplitude_condition:
                        max_second_derivative = second_derivative[local_minimum - 1]
                        best_notch = local_minimum

            # 如果在当前峰值和下一个峰值之间找到了最优的二尖瓣波值点，将它加入列表
            if best_notch is not None:
                dicrotic_notches.append(best_notch)

    return peaks, troughs, dicrotic_notches


# 调用函数并绘制结果
peaks, troughs, dicrotic_notches = find_dicrotic_notches_corrected(ppg_segment)

# 绘制原始信号，并标注峰值和谷值
plt.figure(figsize=(10, 4))
plt.plot(ppg_segment, label='PPG Segment')
plt.scatter(peaks, ppg_segment[peaks], color='green', marker='^', zorder=5, label='Peaks')
plt.scatter(troughs, ppg_segment[troughs], color='orange', marker='v', zorder=5, label='Troughs')
# 标注二尖瓣波值
plt.scatter(dicrotic_notches, ppg_segment[dicrotic_notches], color='red', zorder=5, label='Dicrotic Notches')
plt.title('PPG')
plt.xlabel('index')
plt.ylabel('PPG altitude')
plt.legend()
plt.grid(True)
plt.show()

# 输出二尖瓣波值的索引和对应的PPG幅度值
dicrotic_notches, ppg_segment[dicrotic_notches]

# 计算二尖瓣波值点与其最邻近的谷值的点的垂直距离差值
def calculate_vertical_distances(signal, notches, troughs):
    vertical_distances = []
    for notch in notches:
        # 找到最邻近的谷值
        nearest_trough = min(troughs, key=lambda x: abs(x - notch))
        # 计算垂直距离差值
        vertical_distance = abs(signal[notch] - signal[nearest_trough])
        vertical_distances.append(vertical_distance)
    return vertical_distances

# 计算距离并存储在列表中
vertical_distances = calculate_vertical_distances(ppg_segment, dicrotic_notches, troughs)

# 打印列表内数据
print("二尖瓣波值垂直距离差值列表:", vertical_distances)

# 计算平均值并打印
if len(vertical_distances) == 0:
    average_distance = 0
else:
    average_distance = sum(vertical_distances) / len(vertical_distances)
print("二尖瓣波值平均垂直距离差值:", average_distance)
R1=Altitude/DC_value
R2=average_distance/DC_value
print(R1)
print(R2)


# 以下代码验证二尖瓣波值是否在PPG波形周期中有确定的位置 并使用0.95置信度t检验

# In[2065]:


import matplotlib.pyplot as plt
from scipy import stats
# 计算二尖瓣波值点与周期内第一个波谷的水平距离与整个周期长度的比值
def calculate_ratio(signal, notches, troughs):
    ratios = []

    for notch in notches:
        # 确定在当前二尖瓣波值之前和之后最近的波谷
        previous_troughs = [trough for trough in troughs if trough < notch]
        next_troughs = [trough for trough in troughs if trough > notch]

        if previous_troughs and next_troughs:
            # 找到最近的前一个和后一个波谷
            nearest_previous_trough = max(previous_troughs)
            nearest_next_trough = min(next_troughs)

            # 计算两个波谷之间的水平距离（周期长度）
            cycle_length = nearest_next_trough - nearest_previous_trough

            # 计算二尖瓣波值点与最近前一个波谷之间的水平距离
            horizontal_distance = notch - nearest_previous_trough

            # 计算比值
            if cycle_length != 0:
                ratio = horizontal_distance / cycle_length
                ratios.append(ratio)

    return ratios


# 计算比值
ratios = calculate_ratio(ppg_data, dicrotic_notches, troughs)

# 检验比值是否在一定的置信区间定值
mean_ratio = np.mean(ratios)
std_ratio = np.std(ratios)

# 打印结果
print("二尖瓣与第一个波谷的水平距离与周期长度的比值：", ratios)
print("平均比值：", mean_ratio)
print("标准差：", std_ratio)

# 绘制散点图展示每个周期的比值
plt.figure(figsize=(10, 4))
plt.scatter(range(len(ratios)), ratios, label='Ratio per Cycle')
plt.axhline(y=mean_ratio, color='r', linestyle='-', label=f'Mean Ratio: {mean_ratio:.2f}')
plt.xlabel('Cycle Number')
plt.ylabel('Ratio')
plt.title('Ratio of Horizontal Distance to Cycle Length')
plt.legend()
plt.grid(True)
plt.show()


# 置信度检验
confidence_interval = stats.t.interval(0.95, len(ratios)-1, loc=np.mean(ratios), scale=stats.sem(ratios))

# 打印结果
print("二尖瓣波值与波谷的水平距离比值:", ratios)
print("置信区间:", confidence_interval)


# 使用有二尖瓣波值的二尖瓣波值在波形中占比的置信区间终点重新标注二尖瓣波值

# In[2067]:


import numpy as np
from scipy.signal import find_peaks
def find_dicrotic_notches_by_ratio(signal, peaks, troughs, ratio=0.7):
    notches = []
    for i in range(len(troughs)-1):
        start_trough = troughs[i]
        end_trough = troughs[i+1]
        cycle_length = end_trough - start_trough
        expected_notch_position = start_trough + int(cycle_length * ratio)
        if expected_notch_position not in peaks:
            notches.append(expected_notch_position)
    return notches

#

# 找到峰值和波谷
peaks, _ = find_peaks(ppg_segment, distance=250)
troughs, _ = find_peaks(-ppg_segment, distance=250)

# 找到二尖瓣波值
dicrotic_notches = find_dicrotic_notches_by_ratio(ppg_segment, peaks, troughs)

# 计算垂直距离并打印
vertical_distances = []
for notch in dicrotic_notches:
    nearest_trough_index = np.argmin(np.abs(troughs - notch))
    nearest_trough = troughs[nearest_trough_index]
    vertical_distance = np.abs(ppg_segment[notch] - ppg_segment[nearest_trough])
    vertical_distances.append(vertical_distance)
    print(f"Distance for notch at index {notch}: {vertical_distance}")

# 计算均值
mean_distance = np.mean(vertical_distances)
print(f"Mean vertical distance: {mean_distance}")


# 计算比率并存储
DR = mean_distance / DC_value
print(f"Dicrotic Notch Component to DC Component Ratio: {DR}")
# 绘制原始信号，并标注峰值和谷值
plt.figure(figsize=(10, 4))
plt.plot(ppg_segment, label='PPG Segment')
plt.scatter(peaks, ppg_segment[peaks], color='green', marker='^', zorder=5, label='Peaks')
plt.scatter(troughs, ppg_segment[troughs], color='orange', marker='v', zorder=5, label='Troughs')
# 标注二尖瓣波值
plt.scatter(dicrotic_notches, ppg_segment[dicrotic_notches], color='red', marker='x', zorder=5, label='Dicrotic Notches')
plt.title('PPG Signal with Dicrotic Notches')
plt.xlabel('Index')
plt.ylabel('Amplitude')
plt.legend()
plt.grid(True)
plt.show()


# 计算ppg波的各个线下面积参数 首先PPG波矩形面积（周期内矩形框出的PPG波形面积）与PPG波总线下面积（底边以及PPG波与底边围成的线下面积）绘图并输出面积参数

# In[2068]:


#数据传入
new_peaks_s = peaks
new_troughs_s = troughs
new_dicrotic_notches_s = dicrotic_notches
ppg_s = ppg_segment

def calculate_S1(signal, peaks, troughs):
    S1_areas = []
    cycle_rectangles = []

    # 遍历每对相邻的波谷，计算周期并框出矩形
    for i in range(len(troughs) - 1):
        cycle_start = troughs[i]
        cycle_end = troughs[i + 1]
        cycle_length = cycle_end - cycle_start  # 短边：波谷到波谷的距离

        # 在这个周期内找到最高的波峰值
        peaks_in_cycle = [peak for peak in peaks if cycle_start <= peak < cycle_end]
        if not peaks_in_cycle:  # 如果周期内没有波峰，则跳过
            continue
        max_peak = max(peaks_in_cycle, key=lambda peak: signal[peak])
        amplitude = signal[max_peak] - signal[cycle_start]  # 长边：最大波峰与波谷的垂直距离

        # 计算矩形面积
        area = cycle_length * amplitude
        S1_areas.append(area)

        # 保存矩形的边界以便绘图
        cycle_rectangles.append((cycle_start, cycle_end, signal[cycle_start], signal[max_peak]))

    return S1_areas, cycle_rectangles

# 计算S1面积和周期矩形
S1_areas, cycle_rectangles = calculate_S1(ppg_s, new_peaks_s, new_troughs_s)

# 绘制带有周期矩形的PPG波形，并在矩形内填充颜色

plt.figure(figsize=(12, 6))
plt.plot(ppg_s, label='PPG Segment')
for rect in cycle_rectangles:
    plt.hlines(y=[rect[2], rect[3]], xmin=rect[0], xmax=rect[1], color='grey', alpha=0.5)
    plt.vlines(x=[rect[0], rect[1]], ymin=rect[2], ymax=rect[3], color='grey', alpha=0.5)
    plt.fill_betweenx(y=[rect[2], rect[3]], x1=rect[0], x2=rect[1], color='grey', alpha=0.2)
plt.title('PPG Segment with Filled Rectangles Covering Each Cycle')
plt.xlabel('Sample Index')
plt.ylabel('PPG Amplitude')
plt.legend()
plt.grid(True)
plt.show()

# 输出S1面积
print(S1_areas)
def calculate_ppg_area_under_curve(signal, cycle_rectangles):
    # 计算每个周期内PPG波与底部围成的线下面积
    areas_under_curve = []
    for rect in cycle_rectangles:
        cycle_start = rect[0]
        cycle_end = rect[1]
        baseline = rect[2]  # 波谷值作为基线
        area_under_curve = np.trapz(signal[cycle_start:cycle_end] - baseline, dx=1)
        areas_under_curve.append(area_under_curve)
    return areas_under_curve

# 计算线下面积
areas_under_curve = calculate_ppg_area_under_curve(ppg_s, cycle_rectangles)

# 绘制PPG波形，并填充周期内的线下面积
plt.figure(figsize=(12, 6))
plt.plot(ppg_s, label='PPG Segment')
for rect in cycle_rectangles:
    cycle_start = rect[0]
    cycle_end = rect[1]
    baseline = rect[2]
    plt.fill_between(x=range(cycle_start, cycle_end), y1=ppg_s[cycle_start:cycle_end], y2=baseline, color='blue', alpha=0.2)

plt.title('PPG Segment with Areas Under Curve Filled')
plt.xlabel('Sample Index')
plt.ylabel('PPG Amplitude')
plt.legend()
plt.grid(True)
plt.show()

print(areas_under_curve)


# PPG波二尖瓣波值与距离较远波谷与底边围城线下面积 绘图并输出相应面积 对于框不出面积的周期不求算

# In[2069]:


def calculate_S2_and_plot_corrected(signal, troughs, notches, cycle_rectangles):
    S2_areas = []
    S2_rectangles = []

    # 计算S2面积并保存矩形边界
    for rect in cycle_rectangles:
        cycle_start, cycle_end, _, _ = rect
        # 在周期内找到二尖瓣波值
        notches_in_cycle = [notch for notch in notches if cycle_start <= notch <= cycle_end]
        if not notches_in_cycle:  # 如果周期内没有二尖瓣波值，则跳过
            continue
        notch = notches_in_cycle[0]  # 取周期内第一个二尖瓣波值

        # 在周期内找到与二尖瓣波值最远的波谷
        troughs_in_cycle = [trough for trough in troughs if cycle_start <= trough <= cycle_end]
        if not troughs_in_cycle:  # 如果周期内没有波谷，则跳过
            continue
        farthest_trough = max(troughs_in_cycle, key=lambda trough: abs(trough - notch))
        
        # 确定底边的起始和结束点
        start = min(notch, farthest_trough)
        end = max(notch, farthest_trough)
        trough_level = signal[farthest_trough]  # 底边的垂直位置

        # 计算S2面积
        area = np.trapz(signal[start:end] - trough_level, dx=1)
        S2_areas.append(area)

        # 保存矩形的边界以便绘图
        S2_rectangles.append((start, end, trough_level, signal[notch]))

    # 绘制PPG波形，并填充S2区域
    plt.figure(figsize=(12, 6))
    plt.plot(signal, label='PPG Segment')
    for rect in S2_rectangles:
        plt.fill_between(x=range(rect[0], rect[1]), y1=signal[rect[0]:rect[1]], y2=rect[2], color='lightgreen', alpha=0.3)

    plt.scatter(new_peaks_s, signal[new_peaks_s], color='green', marker='^', zorder=5, label='Peaks')
    plt.scatter(new_troughs_s, signal[new_troughs_s], color='orange', marker='v', zorder=5, label='Troughs')
    plt.scatter(notches, signal[notches], color='red', zorder=5, label='Dicrotic Notches')

    plt.title('PPG Segment with Corrected S2 Areas Under Curve Filled')
    plt.xlabel('Sample Index')
    plt.ylabel('PPG Amplitude')
    plt.legend()
    plt.grid(True)
    plt.show()

    return S2_areas

# 重新计算S2面积并绘制
S2_areas_corrected = calculate_S2_and_plot_corrected(ppg_s, new_troughs_s, new_dicrotic_notches_s, cycle_rectangles)

S2_areas_corrected


#  PPG上升支面积 绘图并输出面积参数

# In[2070]:


def calculate_rising_slope_area_corrected(signal, troughs, peaks):
    # 计算波谷与下一个波峰之间的线下面积
    corrected_rising_slope_areas = []
    corrected_rising_slope_rectangles = []

    for i in range(len(troughs)):
        trough = troughs[i]
        # 找到下一个波峰
        next_peaks = [peak for peak in peaks if peak > trough]
        if not next_peaks:  # 如果没有下一个波峰，则跳过
            continue
        next_peak = next_peaks[0]
        
        # 计算水平距离（底边）
        horizontal_distance = next_peak - trough
        # 垂直位置取波谷的位置
        trough_level = signal[trough]

        # 计算线下面积
        area = np.trapz(signal[trough:next_peak] - trough_level, dx=1)
        corrected_rising_slope_areas.append(area)

        # 保存矩形的边界以便绘图
        corrected_rising_slope_rectangles.append((trough, next_peak, trough_level, signal[next_peak]))

    return corrected_rising_slope_areas, corrected_rising_slope_rectangles

# 计算修正后的上升支的面积
corrected_rising_slope_areas, corrected_rising_slope_rectangles = calculate_rising_slope_area_corrected(ppg_s, new_troughs_s, new_peaks_s)

# 绘制PPG波形，并填充修正后的上升支区域
plt.figure(figsize=(12, 6))
plt.plot(ppg_s, label='PPG Segment')
for rect in corrected_rising_slope_rectangles:
    plt.fill_between(x=range(rect[0], rect[1]), y1=ppg_s[rect[0]:rect[1]], y2=rect[2], color='lightgray', alpha=0.3)
plt.scatter(new_peaks_s, ppg_s[new_peaks_s], color='green', marker='^', zorder=5, label='Peaks')
plt.scatter(new_troughs_s, ppg_s[new_troughs_s], color='orange', marker='v', zorder=5, label='Troughs')
plt.title('PPG Segment with Corrected Rising Slope Areas Filled')
plt.xlabel('Sample Index')
plt.ylabel('PPG Amplitude')
plt.legend()
plt.grid(True)
plt.show()

corrected_rising_slope_areas


# 分别取PPG波各个面积参数的平均值 矩形面积 线下面积 到二尖瓣波值的线下面积 上升支面积

# In[2071]:


# 计算S1_areas, areas_under_curve, S2_areas_corrected, corrected_rising_slope_areas的平均值
average_S1 = sum(S1_areas) / len(S1_areas)
average_areas_under_curve = sum(areas_under_curve)  / len(areas_under_curve)
average_S2_corrected = sum(S2_areas_corrected) / len(S2_areas_corrected)
average_corrected_rising_slope = sum(corrected_rising_slope_areas) / len(corrected_rising_slope_areas)

average_S1, average_areas_under_curve, average_S2_corrected, average_corrected_rising_slope
R1 = average_areas_under_curve/average_S1
a = average_areas_under_curve - average_S2_corrected
R2 = average_S2_corrected / a
R3 = average_corrected_rising_slope / average_S1
R4 = average_S2_corrected / average_S1
print(R1,R2,R3,R4)


# 求算脉搏波动力学参数 上升支垂直距离除以上升支时间 二尖瓣波值垂直距离除以二尖瓣波值对应心脏舒张期时间 二者比值相加 将参数在图中标注

# In[2072]:


def calculate_vertical_horizontal_distances_corrected(signal, troughs, peaks, sampling_rate):
    # 初始化垂直距离和水平距离的列表
    vertical_distances_corrected = []
    horizontal_times_corrected = []  # 水平距离转换为时间

    # 遍历每个波谷和下一个波峰
    for trough in troughs:
        # 找到下一个波峰
        next_peaks = [peak for peak in peaks if peak > trough]
        if not next_peaks:
            continue
        next_peak = next_peaks[0]

        # 计算垂直距离和水平距离
        vertical_distance = signal[next_peak] - signal[trough]
        horizontal_distance = next_peak - trough
        # 将水平距离转换为时间（秒）
        time_duration = horizontal_distance / sampling_rate

        vertical_distances_corrected.append(vertical_distance)
        horizontal_times_corrected.append(time_duration)

    return vertical_distances_corrected, horizontal_times_corrected

# 重新计算垂直和水平距离
vertical_distances_corrected, horizontal_times_corrected = calculate_vertical_horizontal_distances_corrected(ppg_s, new_troughs_s, new_peaks_s, 500)

# 修改代码，以标注周期内波谷到波峰（上升段）的垂直与水平距离
plt.figure(figsize=(12, 6))
plt.plot(ppg_s, label='PPG Segment')
plt.scatter(new_peaks_s, ppg_s[new_peaks_s], color='green', marker='^', zorder=5, label='Peaks')
plt.scatter(new_troughs_s, ppg_s[new_troughs_s], color='orange', marker='v', zorder=5, label='Troughs')
for i in range(min(len(new_troughs_s), len(vertical_distances_corrected))):
    trough = new_troughs_s[i]

    # 确保不会超出 new_peaks_s 的索引范围
    if i >= len(new_peaks_s):
        break
    peak = new_peaks_s[i]

    plt.annotate(f"{vertical_distances_corrected[i]:.2f}", (peak, ppg_s[peak]), textcoords="offset points", xytext=(0,10), ha='center')

    # 检查 horizontal_times_corrected 索引是否有效
    if i < len(horizontal_times_corrected):
        plt.annotate(f"{horizontal_times_corrected[i]:.2f}s", ((trough+peak)/2, ppg_s[trough]), textcoords="offset points", xytext=(0,-15), ha='center')

    # 绘制垂直虚线
    plt.axvline(x=trough, color='grey', linestyle='--', alpha=0.5)
    plt.axvline(x=peak, color='grey', linestyle='--', alpha=0.5)
    # 绘制水平虚线
    plt.hlines(y=ppg_s[trough], xmin=peak, xmax=trough, color='grey', linestyle='--', alpha=0.5)

plt.title('PPG Segment with Marked Vertical and Horizontal Distances (Rising Slope)')
plt.xlabel('Sample Index')
plt.ylabel('PPG Amplitude')
plt.legend()
plt.grid(True)
plt.show()

vertical_distances_corrected, horizontal_times_corrected


# In[2073]:


def calculate_dicrotic_notch_distances(signal, troughs, notches, sampling_rate):
    # 初始化二尖瓣波值的纵坐标和与下一个波谷的水平距离差值
    dicrotic_notch_amplitudes = []
    horizontal_time_differences = []

    # 遍历每个二尖瓣波值
    for notch in notches:
        # 找到下一个波谷
        next_troughs = [trough for trough in troughs if trough > notch]
        if not next_troughs:
            continue
        next_trough = next_troughs[0]

        # 计算二尖瓣波值的纵坐标
        dicrotic_notch_amplitude = signal[notch]
        dicrotic_notch_amplitudes.append(dicrotic_notch_amplitude)

        # 计算水平距离差值并转换为时间
        horizontal_distance = next_trough - notch
        time_difference = horizontal_distance / sampling_rate
        horizontal_time_differences.append(time_difference)

    return dicrotic_notch_amplitudes, horizontal_time_differences

# 计算二尖瓣波值的纵坐标和水平时间差
dicrotic_notch_amplitudes, horizontal_time_differences = calculate_dicrotic_notch_distances(ppg_s, new_troughs_s, new_dicrotic_notches_s, 500)

# 绘制ppg_s的波形并标注
plt.figure(figsize=(12, 6))
plt.plot(ppg_s, label='PPG Segment')
plt.scatter(new_peaks_s, ppg_s[new_peaks_s], color='green', marker='^', zorder=5, label='Peaks')
plt.scatter(new_troughs_s, ppg_s[new_troughs_s], color='orange', marker='v', zorder=5, label='Troughs')
plt.scatter(new_dicrotic_notches_s, ppg_s[new_dicrotic_notches_s], color='red', zorder=5, label='Dicrotic Notches')

# 标注水平时间差和二尖瓣波值的纵坐标
for i, notch in enumerate(new_dicrotic_notches_s):
    if i >= len(new_troughs_s) or i >= len(horizontal_time_differences):
        break
    next_trough = new_troughs_s[i]
    plt.annotate(f"{dicrotic_notch_amplitudes[i]:.2f}", (notch, ppg_s[notch]), textcoords="offset points", xytext=(0,10), ha='center')
    plt.annotate(f"{horizontal_time_differences[i]:.2f}s", ((notch+next_trough)/2, ppg_s[notch]), textcoords="offset points", xytext=(0,-15), ha='center')

    # 画垂直虚线
    plt.axvline(x=notch, color='grey', linestyle='--', alpha=0.5)
    plt.axvline(x=next_trough, color='grey', linestyle='--', alpha=0.5)

plt.title('PPG Segment with Dicrotic Notch Amplitudes and Time Differences')
plt.xlabel('Sample Index')
plt.ylabel('PPG Amplitude')
plt.legend()
plt.grid(True)
plt.show()

dicrotic_notch_amplitudes, horizontal_time_differences


# 求算比值后相加得到脉搏波动力学参数

# In[2074]:


# 使用NumPy进行逐元素的除法
import numpy as np
import pandas as pd
dicrotic_notch_amplitudes_abs = np.abs(dicrotic_notch_amplitudes)
# 确保列表转换为NumPy数组
dicrotic_notch_amplitudes_abs_np = np.array(dicrotic_notch_amplitudes_abs)
horizontal_time_differences_np = np.array(horizontal_time_differences)
vertical_distances_corrected_np = np.array(vertical_distances_corrected)
horizontal_times_corrected_np = np.array(horizontal_times_corrected)

# 确保数组长度相同，以最短数组长度为准
min_length = min(len(dicrotic_notch_amplitudes_abs_np), len(horizontal_time_differences_np))
ratio_amplitude_time = dicrotic_notch_amplitudes_abs_np[:min_length] / horizontal_time_differences_np[:min_length]

min_length = min(len(vertical_distances_corrected_np), len(horizontal_times_corrected_np))
ratio_vertical_horizontal = vertical_distances_corrected_np[:min_length] / horizontal_times_corrected_np[:min_length]

print(ratio_amplitude_time)
print(ratio_vertical_horizontal)
# 计算ratio_amplitude_time和ratio_vertical_horizontal的平均值并相加
average_ratio_amplitude_time = np.mean(ratio_amplitude_time) if ratio_amplitude_time.size > 0 else 0
average_ratio_vertical_horizontal = np.mean(ratio_vertical_horizontal) if ratio_vertical_horizontal.size > 0 else 0

sum_of_averages = average_ratio_amplitude_time + average_ratio_vertical_horizontal
print(sum_of_averages)


# 计算lnDC 表征高阶分量

# In[2075]:


import numpy as np

# 计算DC值的自然对数
log_DC_value = np.log(DC_value)
log_DC_value


# 这是直接将运算结果写入excel表的程序 使用os交互指令 在桌面上新建excel文件 a追加模式

#  

# In[ ]:





# 

# In[2077]:


from openpyxl import load_workbook
import os
from openpyxl import Workbook

# 定义数据字典
data_to_save = {
    "Average Breathing Interval 3.12 (s)": average_breathing_intervalnew,
    "Breathing Period Variability 3.12": breathing_period_variability,
    "Average Heartbeat Interval 3.12 (s)": average_filtered_interval,
    "Heartbeat Period Variability 3.12": filtered_period_variability,
    "R1 AC/DC Component Ratio 3.12": R1,
    "Dicrotic Notch Component to DC Component Ratio 3.12": DR,
    "Average Area of PPG Waveform Bounded by Rectangle 3.12": average_S1,
    "Average Area Under PPG Waveform Curve 3.12": average_areas_under_curve,
    "Average Area Under PPG Waveform Between Dicrotic Notch and First Trough 3.12":average_S2_corrected,
    "Average Area of PPG Waveform Rising Slope 3.12": average_corrected_rising_slope,
    "PPG Pulse Wave Dynamics Parameters 3.12": sum_of_averages,
    "Average DC Component 3.12": DC_value,
    "ln Average DC Component 3.12": log_DC_value,
    "Average Amplitude 3.12": Altitude,
    "Amplitude Variation Period 3.12": period_true,
    "Average Component of Dicrotic Notch 3.12": average_distance,
    "R1":R1,
    "R2":R2,
    "R3":R3,
    "R4":R4
    
}

# 设置文件路径
desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
file_name = "PPG_Data_5.01.xlsx"
file_path = os.path.join(desktop_path, file_name)

# 检查文件是否存在，如果不存在，则创建
if not os.path.exists(file_path):
    wb = Workbook()
    ws = wb.active
    ws.append(list(data_to_save.keys()))
    wb.save(file_path)

# 加载现有的工作簿
wb = load_workbook(file_path)
ws = wb.active

# 追加一行数据
ws.append(list(data_to_save.values()))

# 保存工作簿
wb.save(file_path)

file_path  # 返回文件路径


# 对ppg_filtered进行FFT PSD STFT 小波变换 cmor小波基 短时傅里叶变换分析非平稳信号随时间变化 小波变化提供波形中频率成分所出现在波形中的具体位置 分析波形中瞬态 尖峰 断裂 cmor小波基适合处理震荡类型信号 如PPG波形

# 绘制Poincaré图 显示心率变异性中的周期性和动态变化 图形呈现一个窄的椭圆 说明心率变化较小 心率稳定 

# 将输出图像使用os系统级指令写入桌面figure文件夹中

# In[2078]:


import numpy as np
import matplotlib.pyplot as plt
from scipy.fft import fft
from scipy.signal import welch, stft
import pywt

scales = np.arange(1, 128)
wavelet = 'morl'
duration = 500
ppg_filtered = filtered_ppg

# 创建保存图像的目录
save_dir = "C:\\Users\\HUAWEI\\Desktop\\figure5.01"
if not os.path.exists(save_dir):
    os.makedirs(save_dir)

# FFT分析
N = len(ppg_filtered)
T = 1.0 / 500.0  # 采样间隔，假设为500Hz
xf = np.linspace(0.0, 1.0/(2.0*T), N//2)
yf = fft(ppg_filtered)
plt.figure(figsize=(12, 6))
plt.plot(xf, 2.0/N * np.abs(yf[:N//2]))
plt.title("FFT of PPG Signal")
plt.xlim(0, 15) 
plt.xlabel("Frequency (Hz)")
plt.ylabel("Amplitude")
plt.savefig(os.path.join(save_dir, f"{os.path.splitext(file)[0]}_fft.png"))  # 使用file_name构建文件名
plt.show()  

# 计算并绘制PSD
f, Pxx_den = welch(ppg_filtered, fs=500, nperseg=1024)
plt.figure(figsize=(12, 6))
plt.semilogy(f, Pxx_den)
plt.title("PSD of PPG Signal")
plt.xlabel("Frequency (Hz)")
plt.ylabel("PSD")
plt.savefig(os.path.join(save_dir, f"{os.path.splitext(file)[0]}_psd.png"))  
plt.show()

# 进行STFT
f, t, Zxx = stft(ppg_filtered, fs=500, nperseg=256)
plt.figure(figsize=(12, 6))
plt.pcolormesh(t, f, np.abs(Zxx), shading='gouraud')
plt.title("STFT of PPG Signal")
plt.ylabel("Frequency (Hz)")
plt.xlabel("Time (sec)")
plt.ylim(0, 10)
# 已移除颜色条
plt.savefig(os.path.join(save_dir, f"{os.path.splitext(file)[0]}_stft.png"))  
plt.show()

# 连续小波变换
coefficients, frequencies = pywt.cwt(ppg_filtered, scales, wavelet, 1.0 / sampling_rate)
plt.figure(figsize=(10, 6))
plt.imshow(abs(coefficients), extent=[0, duration, 1, 128], cmap='jet', aspect='auto', 
           vmax=abs(coefficients).max(), vmin=-abs(coefficients).max())
# 已移除颜色条
plt.title('Continuous Wavelet Transform (CWT) of PPG Signal')
plt.ylabel('Scale')
plt.xlabel('Time (sec)')
plt.savefig(os.path.join(save_dir, f"{os.path.splitext(file)[0]}_cwt.png"))  

x = ppg_filtered[:-1]
y = ppg_filtered[1:]
plt.figure(figsize=(6, 6))
plt.scatter(x, y, color='blue')
plt.title('Poincaré Plot')
plt.xlabel('PPG(n)')
plt.ylabel('PPG(n+1)')
plt.savefig(os.path.join(save_dir, f"{os.path.splitext(file)[0]}_poincare.png"))  
plt.show()


# In[ ]:





# In[ ]:





# 

# In[ ]:





# 

# In[ ]:





# 

# In[ ]:





# 

# In[ ]:





# 

# In[ ]:





# 

# In[ ]:





# 

# 

# In[ ]:





# In[ ]:





# In[ ]:




